# pip install openpyxl
from openpyxl import load_workbook
# xlsxwriter는 데이터 write만 가능했다
# openpyxl은 read&write 둘 다 가능하다
# 셀 참조 방식도 예쁘지만 속도가 인지도 있는 엑셀 라이브러리들 중 가장 느리다

wb = load_workbook('test.xlsx')
sheet = wb['test_sheet']

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A1'].value)
sheet['A2'] = 'New Data'

wb.save('test.xlsx')
