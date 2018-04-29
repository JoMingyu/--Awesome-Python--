# https://github.com/ericgazoni/openpyxl
# pip install openpyxl
from openpyxl import load_workbook, Workbook
# read&write 둘 다 가능한 엑셀 라이브러리
# 셀 참조 방식도 간단하고 직관적이지만 인지도 있는 엑셀 라이브러리들 중 속도가 가장 느리다
# 속도에 구애받지 않는 환경이라면 가장 편하다

wb = Workbook()
# workbook 생성

ws = wb.active
ws.title = 'sheet 1'
# worksheet 활성화 후 이름 변경

ws2 = wb.create_sheet('sheet 2')
# create_sheet 메소드를 사용할 수도 있음

ws['A1'] = 42
# 셀에 직접적으로 값을 할당해줄 수 있음

wb.save('test.xlsx')
# 파일로 저장

# --

wb = load_workbook('test.xlsx')
# 이미 존재하는 workbook load

sheet = wb['sheet 1']
# worksheet에 접근

print(sheet.max_row)
print(sheet.max_column)
# 최대 row와 최대 column

print(sheet['A1'].value)
# 셀의 value에 접근
sheet['A2'] = 'New Data'
# 셀의 데이터 수정

wb.save('test.xlsx')
